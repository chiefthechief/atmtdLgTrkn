import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# --- Configuration ---
EXCEL_FILE = r"{your file location}"
HEADERS = ["Date", "Student/Staff Number", "Issue Reported", "Solution Applied", "State"]

# --- Predefined lists (customize these as needed) ---
COMMON_ISSUES = [
    #enter your common issues here
]

COMMON_SOLUTIONS = [
    # enter your common solutions herer
]

# --- Helper function to display a numbered list and get selection ---
def pick_from_list(prompt, options):
    """Show options with numbers, let user pick a number or enter custom text.
    Returns the chosen text, or None if user quits."""
    print(f"\n{prompt}")
    for i, opt in enumerate(options, 1):
        print(f"  {i}. {opt}")
    print("  Or type your own text (or 'quit' to exit).")
    
    while True:
        choice = input("Your choice (number or text): ").strip()
        if choice.lower() == "quit":
            return None
        # Try to interpret as number
        try:
            num = int(choice)
            if 1 <= num <= len(options):
                return options[num-1]
            else:
                print(f"Please enter a number between 1 and {len(options)}.")
        except ValueError:
            # Not a number – treat as custom text
            if choice:
                return choice
            else:
                print("Input cannot be empty.")

# --- Ensure directory and load/create workbook ---
os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)

if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)

print("=== IT Support Log Entry Tool ===")
print("Type 'quit' at any prompt to exit.\n")

while True:
    print("\n--- New Entry ---")
    
    # Date is automatic
    date_today = datetime.now().strftime("%m/%d/%Y")
    print(f"Date: {date_today}")
    
    # Student number
    student_num = input("Student/Staff Number (or 'quit'): ").strip()
    if student_num.lower() == "quit":
        break
    if not student_num:
        print("Student number cannot be empty.")
        continue
    
    # Issue – pick from list or custom
    issue = pick_from_list("Select Issue Reported:", COMMON_ISSUES)
    if issue is None:   # user quit
        break
    
    # Solution – pick from list or custom
    solution = pick_from_list("Select Solution Applied:", COMMON_SOLUTIONS)
    if solution is None:
        break
    
    # State with validation
    state = ""
    while state.lower() not in ["resolved", "pending"]:
        state_input = input("State (resolved/pending) or 'quit': ").strip().lower()
        if state_input == "quit":
            break
        if state_input in ["resolved", "pending"]:
            state = state_input
        else:
            print("Please enter either 'resolved' or 'pending'.")
    if state == "":
        break
    
    # Append row
    new_row = [date_today, student_num, issue, solution, state.capitalize()]
    ws.append(new_row)
    
    # Apply dropdown for State column (E) to all data rows
    dv = DataValidation(type="list", formula1='"Resolved,Pending"', allow_blank=True)
    dv.add(f"E2:E{ws.max_row}")
    ws.add_data_validation(dv)
    
    # Save
    try:
        wb.save(EXCEL_FILE)
        print("✅ Entry saved.")
    except PermissionError:
        print("\n❌ Could not save. Please close the Excel file and try again.")
        break

print("\nExiting. Goodbye!")

